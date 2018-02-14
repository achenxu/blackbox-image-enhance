import sys
import os
import os.path
import json
import numpy as np
import cv2
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

ex_str = [
    "removed string what you want",
    "blabla"
]

cafe_base = os.environ['CAFE_BASE_URL']


def exclude_template(body):
    for s in ex_str:
        body = body.replace(s, '')

    body = re.sub('\s+', ' ', body)
    body = body.replace('( )', '')
    body = body.replace('^^', '')
    return body


def append_data(num, ws, line):
    mp4_file = './data_mp4/%d-%d.mp4' % (num, 0)
    json_file = './data_json/%d-%d.json' % (num, 0)
    html_file = './data_html/%d.html' % (num)
    sl = str(line)

    if not os.path.exists(mp4_file):
        return None

    with open(json_file, 'r') as f:
        j = json.loads(f.read())

    with open(html_file, 'r') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    print ("Add %d artichle" % num)
    ws['C' + sl].hyperlink = '%s/%d' % (cafe_base, num)
    ws['C' + sl].value = num

    ws['D' + sl].value = j['meta']['subject']

    tbody = soup.select('.tbody')
    body = tbody[0].text.strip() if len(tbody) >= 1 else ""
    ws['E' + sl].value = exclude_template(body)
    reply = [x.text.strip() for x in soup.select('.box-reply2 .comm')]
    ws['F' + sl].value = '\n'.join(reply)

    tdate = soup.select('.tit-box .m-tcol-c.date')
    sdate = tdate[0].text.strip() if len(tdate) >= 1 else ""
    ws['G' + sl].value = sdate

    videos = j['videos']['list']
    seq = [float(x['bitrate']['video']) for x in videos]
    max_value = max(seq)
    data = [(i, val) for (i, val) in enumerate(videos)
            if float(val['bitrate']['video']) == max_value]

    ws['H' + sl].value = data[0][1]['encodingOption']['width']
    ws['I' + sl].value = data[0][1]['encodingOption']['height']
    ws['J' + sl].value = data[0][1]['encodingOption']['name']
    ws['K' + sl].value = data[0][1]['bitrate']['video']
    ws['L' + sl].value = mp4_file

    # get imagePath
    vidcap = cv2.VideoCapture(mp4_file)
    frames = np.linspace(0, vidcap.get(cv2.CAP_PROP_FRAME_COUNT) - 1, 1,
                         dtype=np.int)
    col = 13
    cell_width = 17
    ws.row_dimensions[line].height = 100
    for f in frames:
        tname = '%d-%d.jpg' % (num, f)

        vidcap.set(cv2.CAP_PROP_POS_FRAMES, f)
        success, image = vidcap.read()

        if not success:
            print(" Bad Image")
            col += 1
            continue

        height, width, layers = image.shape
        ratio = width / 128
        new_h = int(height / ratio)
        new_w = int(width / ratio)
        resize = cv2.resize(image, (new_w, new_h))

        cv2.imwrite(tname, resize)
        ws.column_dimensions[get_column_letter(col)].width = cell_width
        img = Image(tname, size=(width * 9, 10000))
        c = ws[get_column_letter(col) + sl]
        img.anchor(c, 'oneCell')
        ws.add_image(img)
        # os.remove(tname)
        col += 1

    return ws


if __name__ == '__main__':
    wb = load_workbook(filename='excel-template.xlsx')
    ws = wb['data']

    start = int(sys.argv[1])
    end = int(sys.argv[2])
    step = 1 if end > start else -1

    line = 4
    for i in range(start, end, step):
        if append_data(i, ws, line):
            line += 1

    wb.save('test.xlsx')
